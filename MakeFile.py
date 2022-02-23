import os
from openpyxl import load_workbook, Workbook
import datetime, locale

data_file = 'SOP.xlsx'

# Load the entire workbook.
wb = load_workbook(data_file)

#Get month name in Polish
locale.setlocale(locale.LC_ALL, 'pl_PL')
data = datetime.datetime.now()
data_out = (data.strftime("%B"))
data_filename = (data.strftime("%d %B %Y"))

# Load one worksheet.
ws = wb[data_out.capitalize()]
all_rows = list(ws.rows)

#Create new workbook
wd = Workbook()
dest_filename = str(data_filename)+".xlsx"
wa = wd.active
wss = wd.create_sheet(data_out,0)
other_rows = list(wa.rows)

matrix = []
helper = 0

#Insert data to array
for row in all_rows[1:400]:
        b = '-'
        state = row[27].value
        if(state is not None and state is not b ):
            helper+=1
            matrix.append(state)
            #print(f"{state}")

for row in all_rows[1:400]:
        b = '-'
        state = row[30].value
        if(state is not None and state is not b ):
            helper+=1
            matrix.append(state)
           

for row in all_rows[1:400]:
        b = '-'
        state = row[33].value 
        if(state is not None and state is not b ):
            helper+=1
            matrix.append(state)
            
wss.insert_rows(helper)

#Insert vaule from array to cells
row_i = 1
for i in matrix:
    wss.cell(row=row_i, column=1).value = i
    row_i+=1


wd.save(filename=dest_filename)

if os.path.exists('SOP.xlsx'):
    os.remove('SOP.xlsx')
